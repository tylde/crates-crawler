from openpyxl import load_workbook, Workbook

from config.index import SELL_ORDERS_SHEET_NAME, PRICE_SHEET_NAME
from crates_crawler.model.Crate import Crate
from crates_crawler.model.OrdersHistogramData import OrdersHistogramData
from crates_crawler.model.PriceOverviewData import PriceOverviewData
from crates_crawler.spreadsheet.sheet.ColorSheet import ColorSheet
from crates_crawler.spreadsheet.sheet.HistogramSheet import HistogramSheet
from crates_crawler.spreadsheet.sheet.PriceSheet import PriceSheet
from crates_crawler.spreadsheet.sheet.SellOrdersSheet import SellOrdersSheet
from crates_crawler.utils.Time import Time


class Spreadsheet:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = self._load_workbook()

    def _load_workbook(self):
        try:
            stopper = Time()
            print("Loading spreadsheet...")
            stopper.start()
            workbook = load_workbook(self.filename)
            stopper.end()
            print(f"Done. ({stopper.result:0.3f}s)")
            return workbook
        except FileNotFoundError:
            stopper = Time()
            print("Creating spreadsheet...")
            stopper.start()
            workbook = self._create_workbook()
            stopper.end()
            print(f"Done. ({stopper.result:0.3f}s)")
            return workbook

    def _create_workbook(self):
        workbook = Workbook()
        workbook[workbook.sheetnames[0]].title = SELL_ORDERS_SHEET_NAME
        return workbook

    def insert_price_overview_data(self, crate: Crate, time, data: PriceOverviewData):
        histogram_sheet = HistogramSheet(self.workbook, crate)
        histogram_sheet.insert_price_overview_data(time, data)

    def insert_histogram_data(self, crate: Crate, datetime, data: OrdersHistogramData):
        sell_orders_sheet = SellOrdersSheet(self.workbook, SELL_ORDERS_SHEET_NAME)
        sell_orders_sheet.insert_histogram_data(datetime, data, crate)

        price_sheet = PriceSheet(self.workbook, PRICE_SHEET_NAME)
        price_sheet.insert_histogram_data(datetime, data, crate)

        histogram_sheet = HistogramSheet(self.workbook, crate)
        histogram_sheet.insert_histogram_data(datetime, data)

    def insert_colors(self):
        colors_sheet = ColorSheet(self.workbook, 'Colors')
        colors_sheet.fill_colors()

    def save(self):
        self.workbook.save(self.filename)

    def __str__(self):
        return '<Spreadsheet \"' + self.filename + '\">'
