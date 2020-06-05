from openpyxl.utils import get_column_letter

from config.sheet_config import DATE_COLUMN_INDEX, STATUS_COLUMN_INDEX
from crates_crawler.spreadsheet.cell.Cell import Cell
from crates_crawler.utils.Time import Time


class Sheet:
    def __init__(self, workbook, name):
        self.workbook = workbook
        self.name = name

        self.sheet = self._get_worksheet()

    def _get_worksheet(self):
        try:
            index = self.workbook.sheetnames.index(self.name)
            return self.workbook.worksheets[index]
        except ValueError:
            self._create()
            index = self.workbook.sheetnames.index(self.name)
            return self.workbook.worksheets[index]

    def _create(self):
        self.workbook.create_sheet(self.name)

    def _get_index(self):
        try:
            return self.workbook.sheetnames.index(self.name)
        except ValueError:
            self._create()
            return self.workbook.sheetnames.index(self.name)

    def _set_active(self):
        sheet_index = self._get_index()
        self.workbook.active = sheet_index

    def cell(self, cell_name):
        return Cell.get_by_name(self.workbook, self.name, cell_name)

    def cell_by_index(self, row, column):
        return Cell.get_by_index(self.workbook, self.name, row, column)

    def set_column_width(self, column: str, width):
        self.sheet.column_dimensions[column].width = width

    def set_column_index_width(self, index: int, width):
        self.sheet.column_dimensions[get_column_letter(index)].width = width

    def insert_column(self, index):
        self.sheet.insert_cols(index, 1)

    def _find_by_date_row_index(self, time):
        for cell in self.sheet[get_column_letter(DATE_COLUMN_INDEX)]:
            if cell.value == time:
                return cell.row
        return -1

    def _find_available_row_index(self):
        for cell in self.sheet[get_column_letter(DATE_COLUMN_INDEX)]:
            if cell.value is None:
                return cell.row
        else:
            return cell.row + 1

    def _find_by_header_name(self, name):
        for cell in self.sheet[1]:
            if cell.value == name:
                return cell.column
        return -1

    def _find_available_column_index(self):
        for cell in self.sheet[1]:
            if cell.value is None:
                return cell.column
        else:
            return cell.column + 1

    def _is_row_valid(self, row_index):
        status_cell = self.cell_by_index(row_index, STATUS_COLUMN_INDEX)
        if status_cell.value is None:
            return True
        return False

    def _has_row_date(self, row_index):
        date_cell = self.cell_by_index(row_index, DATE_COLUMN_INDEX)
        if date_cell.value is not None:
            return True
        return False

    def _get_row_datetime(self, row_index):
        return self.cell_by_index(row_index, DATE_COLUMN_INDEX).value

    def _get_last_day_row_index(self, reference_index):
        reference_date = self._get_row_datetime(reference_index)
        reference_date_timestamp = Time.timestamp(reference_date)
        current_index = reference_index - 1
        result_diff = 24 * 3600
        result_index = -1
        diff = 24 * 3600

        while current_index > 0 and diff > (-2 * 3600):
            if self._has_row_date(current_index) is False or self._is_row_valid(current_index) is False:
                current_index -= 1
                continue

            current_date = self._get_row_datetime(current_index)
            current_date_timestamp = Time.timestamp(current_date)

            diff = current_date_timestamp - (reference_date_timestamp - 24 * 3600)
            diff_abs = abs(diff)
            if diff_abs < result_diff and diff_abs < (2 * 3600):
                result_diff = diff_abs
                result_index = current_index

            current_index -= 1

        return result_index

    def _get_ratio_color(self, ratio, invert=False):
        if invert is False:
            diff = ratio - 1
        else:
            diff = -1 * (ratio - 1)

        pattern_name = "DARK"
        if -0.005 < diff < 0.005:
            pattern_name = "DARK"
        elif diff > 0.005:
            pattern_name = "GREEN"
        else:
            pattern_name = "RED"

        diff_abs = abs(diff)

        level = 0
        if diff_abs < 0.025:
            level = 0
        elif 0.025 <= diff_abs < 0.05:
            level = 1
        elif 0.05 <= diff_abs < 0.075:
            level = 2
        elif 0.075 <= diff_abs < 0.1:
            level = 3
        elif 0.1 <= diff_abs < 0.25:
            level = 4
        else:
            level = 5

        return pattern_name, level
