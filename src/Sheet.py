from openpyxl.utils import get_column_letter

from src.Cell import Cell

from config.sheet_config import DATE_COLUMN


class Sheet:
    def __init__(self, workbook, name):
        self.workbook = workbook
        self.name = name

        self.sheet = self._get_worksheet()

    def _get_worksheet(self):
        try:
            index = self.workbook.sheetnames.index(self.name)
            return self.workbook.worksheets[index]
        except ValueError:
            self._create()
            index = self.workbook.sheetnames.index(self.name)
            return self.workbook.worksheets[index]

    def _create(self):
        self.workbook.create_sheet(self.name)

    def _get_index(self):
        try:
            return self.workbook.sheetnames.index(self.name)
        except ValueError:
            self._create()
            return self.workbook.sheetnames.index(self.name)

    def _set_active(self):
        sheet_index = self._get_index()
        self.workbook.active = sheet_index

    def cell(self, cell_name):
        return Cell.get_by_name(self.workbook, self.name, cell_name)

    def cell_by_index(self, row, column):
        return Cell.get_by_index(self.workbook, self.name, row, column)

    def set_column_width(self, column: str, width):
        self.sheet.column_dimensions[column].width = width

    def set_column_index_width(self, index: int, width):
        self.sheet.column_dimensions[get_column_letter(index)].width = width

    def insert_column(self, index):
        self.sheet.insert_cols(index, 1)

    def _find_by_date_row_index(self, time):
        for cell in self.sheet[DATE_COLUMN]:
            if cell.value == time:
                return cell.row
        return -1

    def _find_available_row_index(self):
        for cell in self.sheet[DATE_COLUMN]:
            if cell.value is None:
                return cell.row
        else:
            return cell.row + 1
