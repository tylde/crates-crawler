from openpyxl import load_workbook, Workbook

from src.ColorSheet import ColorSheet
from src.Crate import Crate
from src.HistogramSheet import HistogramSheet
from src.OrdersHistogramData import OrdersHistogramData
from src.OverviewSheet import OverviewSheet
from src.PriceOverviewData import PriceOverviewData

from config.index import OVERVIEW_SHEET_NAME


class Spreadsheet:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = self._load_workbook()

    def _load_workbook(self):
        try:
            return load_workbook(self.filename)
        except FileNotFoundError:
            return self._create_workbook()

    def _create_workbook(self):
        workbook = Workbook()
        workbook[workbook.sheetnames[0]].title = OVERVIEW_SHEET_NAME
        workbook.save(self.filename)
        return workbook

    def insert_price_overview_data(self, crate: Crate, time, data: PriceOverviewData):
        histogram_sheet = HistogramSheet(self.workbook, crate)
        histogram_sheet.insert_price_overview_data(time, data)

    def insert_histogram_data(self, crate: Crate, datetime, data: OrdersHistogramData):
        histogram_sheet = HistogramSheet(self.workbook, crate)
        histogram_sheet.insert_histogram_data(datetime, data)

        overview_sheet = OverviewSheet(self.workbook, OVERVIEW_SHEET_NAME)
        overview_sheet.insert_histogram_data(datetime, data, crate)

    def insert_colors(self):
        colors_sheet = ColorSheet(self.workbook, 'Colors')
        colors_sheet.fill_colors()

    def save(self):
        self.workbook.save(self.filename)


    def __str__(self):
        return '<Spreadsheet \"' + self.filename + '\">'
