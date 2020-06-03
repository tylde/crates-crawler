from openpyxl.styles import Protection, numbers

from src.CellAlignment import CellAlignment
from src.CellBorder import CellBorder
from src.CellFill import CellFill
from src.CellFont import CellFont


class Cell(CellAlignment, CellBorder, CellFill, CellFont):
    def __init__(self, cell):
        CellAlignment.__init__(self, cell)
        CellBorder.__init__(self, cell)
        CellFont.__init__(self, cell)

        self.cell = cell

    @classmethod
    def get_by_name(cls, workbook, sheet_name, cell_name):
        sheet_index = workbook.sheetnames.index(sheet_name)
        cell = workbook.worksheets[sheet_index][cell_name]
        return cls(cell)

    @classmethod
    def get_by_index(cls, workbook, sheet_name, row, column):
        sheet_index = workbook.sheetnames.index(sheet_name)
        cell = workbook.worksheets[sheet_index].cell(row, column)
        return cls(cell)

    @property
    def value(self):
        return self.cell.value

    def set_value(self, value):
        self.cell.value = value
        return self

    @property
    def row(self):
        return self.cell.row

    @property
    def column(self):
        return self.cell.column

    @property
    def coordinate(self):
        return self.cell.coordinate

    def number_format(self):
        self.cell.number_format = numbers.BUILTIN_FORMATS[3]
        return self

    def lock(self):
        self.cell.protection = Protection(locked=True)
        return self

    def unlock(self):
        self.cell.protection = Protection(locked=False)
        return self

    def __str__(self):
        return '<CellWrapper[' + str(self.row) + ',' + str(self.column) + '] \"' + self.coordinate + '\">'
